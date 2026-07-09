from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
import json
from app.models import Order, Product, Customer
from app.decorators import staff_required


@login_required
def report_views(request):
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders = Order.objects.all()
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    # Summary
    total_revenue = orders.filter(status='delivered').aggregate(
        total=Sum('total_amount'))['total'] or 0
    total_orders = orders.count()
    delivered = orders.filter(status='delivered').count()
    pending = orders.filter(status='pending').count()
    cancelled = orders.filter(status='cancelled').count()

    # Top products by revenue
    from app.models import OrderItem
    top_products = OrderItem.objects.values(
        'product__name'
    ).annotate(
        revenue=Sum('unit_price'),
        qty=Sum('quantity')
    ).order_by('-revenue')[:10]

    # Top customers
    top_customers = Customer.objects.annotate(
        order_count=Count('orders'),
        total_spent=Sum('orders__total_amount')
    ).order_by('-total_spent')[:10]

    # Monthly revenue (last 12 months)
    today = timezone.now().date()
    monthly_data = []
    for i in range(11, -1, -1):
        month = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        rev = Order.objects.filter(
            status='delivered',
            created_at__year=month.year,
            created_at__month=month.month
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        monthly_data.append({'month': month.strftime('%b %Y'), 'revenue': float(rev)})

    context = {
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'delivered': delivered,
        'pending': pending,
        'cancelled': cancelled,
        'top_products': top_products,
        'top_customers': top_customers,
        'monthly_data': json.dumps(monthly_data),
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'report.html', context)


@login_required
def chart_data(request):
    today = timezone.now().date()
    # Last 30 days daily revenue
    daily = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        rev = Order.objects.filter(
            status='delivered', created_at__date=day
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        daily.append({'date': day.strftime('%m/%d'), 'revenue': float(rev)})

    # Status distribution
    status_data = list(Order.objects.values('status').annotate(count=Count('id')))

    return JsonResponse({'daily': daily, 'status': status_data})


@staff_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    wb = openpyxl.Workbook()

    # ─── Orders Sheet ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Orders"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a56db")

    headers = ['Order #', 'Customer', 'Status', 'Payment', 'Total Amount', 'Date']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws1.column_dimensions[get_column_letter(col)].width = 18

    orders = Order.objects.select_related('customer').all()
    for row, order in enumerate(orders, 2):
        ws1.cell(row=row, column=1, value=order.order_number)
        ws1.cell(row=row, column=2, value=order.customer.full_name)
        ws1.cell(row=row, column=3, value=order.get_status_display())
        ws1.cell(row=row, column=4, value=order.get_payment_method_display())
        ws1.cell(row=row, column=5, value=float(order.total_amount))
        ws1.cell(row=row, column=6, value=order.created_at.strftime('%Y-%m-%d'))

    # ─── Products Sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Products")
    p_headers = ['Name', 'SKU', 'Category', 'Price', 'Stock', 'Active']
    for col, h in enumerate(p_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col)].width = 18

    for row, p in enumerate(Product.objects.select_related('category').all(), 2):
        ws2.cell(row=row, column=1, value=p.name)
        ws2.cell(row=row, column=2, value=p.sku)
        ws2.cell(row=row, column=3, value=p.category.name if p.category else '')
        ws2.cell(row=row, column=4, value=float(p.price))
        ws2.cell(row=row, column=5, value=p.stock)
        ws2.cell(row=row, column=6, value='Yes' if p.is_active else 'No')

    # ─── Customers Sheet ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Customers")
    c_headers = ['Name', 'Email', 'Phone', 'City', 'Country', 'Total Orders']
    for col, h in enumerate(c_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[get_column_letter(col)].width = 20

    for row, c in enumerate(Customer.objects.all(), 2):
        ws3.cell(row=row, column=1, value=c.full_name)
        ws3.cell(row=row, column=2, value=c.email)
        ws3.cell(row=row, column=3, value=c.phone)
        ws3.cell(row=row, column=4, value=c.city)
        ws3.cell(row=row, column=5, value=c.country)
        ws3.cell(row=row, column=6, value=c.total_orders())

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)
    return response


@staff_required
def export_pdf(request):
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
    except ImportError:
        return HttpResponse("reportlab not installed. Run: pip install reportlab", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=20)
    elements.append(Paragraph("Sales Performance & Order Management Report", title_style))
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary stats
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='delivered').aggregate(
        total=Sum('total_amount'))['total'] or 0
    total_customers = Customer.objects.count()

    elements.append(Paragraph("Summary", styles['Heading2']))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Orders', str(total_orders)],
        ['Total Revenue', f"${float(total_revenue):,.2f}"],
        ['Total Customers', str(total_customers)],
        ['Total Products', str(Product.objects.count())],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Orders table
    elements.append(Paragraph("Recent Orders", styles['Heading2']))
    orders = Order.objects.select_related('customer').order_by('-created_at')[:20]
    order_data = [['Order #', 'Customer', 'Status', 'Payment', 'Amount', 'Date']]
    for o in orders:
        order_data.append([
            o.order_number,
            o.customer.full_name,
            o.get_status_display(),
            o.get_payment_method_display(),
            f"${float(o.total_amount):,.2f}",
            o.created_at.strftime('%Y-%m-%d'),
        ])

    col_widths = [1.5*inch, 2*inch, 1.2*inch, 1.5*inch, 1.2*inch, 1.2*inch]
    order_table = Table(order_data, colWidths=col_widths)
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(order_table)

    doc.build(elements)
    return response